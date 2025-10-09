# app.py
# ==============================================================
# ⚓ Naval Machinery Health Monitoring Web App (Complete)
# ==============================================================
# - Rule-based + RandomForest ML modes
# - Many visualizations + PDF report + maintenance log
# - Robust column detection & encoding handling
# ==============================================================

import streamlit as st
# 1) set_page_config must be the first streamlit call
st.set_page_config(page_title="Naval Machinery Health Monitoring",
                   page_icon="⚓", layout="wide")

import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import io, os, re, pickle, tempfile, uuid
from datetime import datetime
from pathlib import Path
from fpdf import FPDF
import chardet
import sys
from scipy.stats import skew, kurtosis

# allow local imports from src
sys.path.append(os.path.join(os.path.dirname(__file__), "src"))

# ---------------------------
# Paths & config
# ---------------------------
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
LOG_PATH = DATA_DIR / "maintenance_log.csv"
MODEL_PATH = Path("src/rf_model.pkl")
FONTS_DIR = Path("fonts")
LOCAL_DEJAVU = FONTS_DIR / "DejaVuSans.ttf"
SYSTEM_DEJAVU = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"

# ---------------------------
# Page header & About
# ---------------------------
st.title("⚓ Equipment Health Condition Monitoring for Naval Ships")
st.write("Predictive maintenance prototype — rule-based + machine learning. ")
with st.sidebar:
    st.header("About")
    st.markdown("**Developed by:** Dr. Awujoola Olalekan")
    st.markdown("This tool supports manual CSV/XLSX inputs and (optionally) a pre-trained RandomForest model saved at `src/rf_model.pkl` for ML predictions.")
    st.markdown("---")

# ---------------------------
# Sidebar: Upload / Mode / Example
# ---------------------------
st.sidebar.header("Upload & Options")
uploaded = st.sidebar.file_uploader("Upload dataset (CSV or Excel)", type=["csv", "xlsx"])
mode = st.sidebar.radio("Evaluation mode", ["Automatic Rule-based", "Machine Learning (Random Forest)", "Compare both"])
st.sidebar.markdown("---")
if st.sidebar.button("Download example dataset"):
    example = pd.DataFrame({
        "SYSTEM": ["Propulsion", "Cooling", "Auxiliary"],
        "EQUIPMENT": ["Main Engine (Port)", "Fresh Water Pump", "Generator"],
        "HEALTH INDICATOR (HI)": ["Lube oil pressure (bar)", "Oil temp (°C)", "Voltage (V)"],
        "SYNTHETIC VALUE": [4.3, 78.0, 440.0],
        "MIN–MAX THRESHOLDS": ["3.5 – 6.0", "60 – 90", "420 – 450"],
        "WORKING VALUE ONBOARD": ["3.6 bar", "82 °C", "438 V"],
        "REMARKS": ["Nominal", "Slightly high", "Nominal"],
        "Actual_Status": ["Healthy", "Warning", "Healthy"]
    })
    st.sidebar.download_button("Download example.csv", example.to_csv(index=False).encode("utf-8"), "example.csv", "text/csv")

st.sidebar.markdown("---")
st.sidebar.info("If ML mode selected, the app will attempt to load model bundle from `src/rf_model.pkl`")

# ---------------------------
# Load model bundle if present
# ---------------------------
model_bundle = None
if MODEL_PATH.exists():
    try:
        with open(MODEL_PATH, "rb") as f:
            model_bundle = pickle.load(f)
    except Exception as e:
        st.sidebar.error(f"Could not load ML bundle: {e}")
else:
    st.sidebar.warning("No ML model found at src/rf_model.pkl — ML modes disabled until model is added.")

# ---------------------------
# Helper utilities
# ---------------------------
def detect_encoding_bytes(b: bytes):
    try:
        enc = chardet.detect(b[:200000])["encoding"]
        return enc or "utf-8"
    except Exception:
        return "utf-8"

def sanitize_columns(df: pd.DataFrame):
    cols = []
    for c in df.columns:
        s = str(c).strip()
        s = s.replace("\u2013", "-").replace("\u2014", "-").replace("\xa0", " ")
        s = re.sub(r"\s+", " ", s).strip()
        s = s.replace(" ", "_").replace("-", "_")
        s = re.sub(r"[^0-9A-Za-z_()]", "", s)
        cols.append(s)
    df = df.copy()
    df.columns = cols
    return df

def find_column(df, candidates):
    # tries exact match, normalized, contains
    norm_map = {re.sub(r'[^0-9a-z]', '', c.lower()): c for c in df.columns}
    for cand in candidates:
        k = re.sub(r'[^0-9a-z]', '', cand.lower())
        if k in norm_map:
            return norm_map[k]
    # contains
    for col in df.columns:
        low = re.sub(r'[^0-9a-z]', '', col.lower())
        for cand in candidates:
            k = re.sub(r'[^0-9a-z]', '', cand.lower())
            if k and k in low:
                return col
    return None

def extract_min_max(s):
    if pd.isna(s):
        return (np.nan, np.nan)
    nums = re.findall(r"[-+]?\d*\.\d+|\d+", str(s))
    if len(nums) >= 2:
        return float(nums[0]), float(nums[1])
    elif len(nums) == 1:
        return float(nums[0]), np.nan
    return (np.nan, np.nan)

def extract_numeric(s):
    if pd.isna(s):
        return np.nan
    m = re.search(r"[-+]?\d*\.\d+|\d+", str(s))
    return float(m.group(0)) if m else np.nan

def compute_hi_rule(val, lo, hi):
    if np.isnan(val) or np.isnan(lo) or np.isnan(hi):
        return np.nan
    if lo <= val <= hi:
        return 1.0
    elif (lo * 0.9) <= val <= (hi * 1.1):
        return 0.5
    return 0.0

def save_fig_to_buffer(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf

# PDF generation helper
def add_dejavu_font(pdf_obj):
    try:
        if LOCAL_DEJAVU.exists():
            pdf_obj.add_font("DejaVu", "", str(LOCAL_DEJAVU), uni=True)
            return "DejaVu"
        if Path(SYSTEM_DEJAVU).exists():
            pdf_obj.add_font("DejaVu", "", SYSTEM_DEJAVU, uni=True)
            return "DejaVu"
    except Exception:
        pass
    return None

def generate_pdf_report(title, metrics_df, interpretation, suggestions, chart_buffers, authors="Navy Capt. Daya Abdullahi & Dr. Awujoola Olalekan J"):
    # Writes a PDF to a temporary file and returns bytes
    try:
        tmpname = Path(tempfile.gettempdir()) / f"naval_report_{uuid.uuid4().hex}.pdf"
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)
        fnt = add_dejavu_font(pdf)
        if fnt:
            pdf.set_font(fnt, size=14)
        else:
            pdf.set_font("Arial", size=14)

        # Cover page
        pdf.add_page()
        pdf.cell(0, 10, title, ln=True, align="C")
        pdf.ln(4)
        if fnt:
            pdf.set_font(fnt, size=12)
        else:
            pdf.set_font("Arial", size=12)
        pdf.cell(0, 8, "Predictive Maintenance Model Summary", ln=True, align="C")
        pdf.ln(6)
        pdf.multi_cell(0, 7, f"Authors: {authors}")
        pdf.multi_cell(0, 7, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        pdf.ln(6)
        pdf.multi_cell(0, 7, "Confidential - For Maintenance Use Only")

        # Metrics page
        pdf.add_page()
        pdf.set_font("Arial", size=11) if not fnt else pdf.set_font(fnt, size=11)
        pdf.cell(0, 8, "Model Performance Metrics", ln=True)
        pdf.ln(4)
        for _, r in metrics_df.iterrows():
            pdf.cell(0, 7, f"{r['Metric']}: {r['Value']}", ln=True)
        pdf.ln(6)
        pdf.multi_cell(0, 7, "Interpretation:")
        pdf.multi_cell(0, 7, interpretation)
        pdf.ln(4)
        pdf.multi_cell(0, 7, "Suggested Maintenance Actions:")
        if isinstance(suggestions, (list, tuple)):
            for s in suggestions:
                pdf.multi_cell(0, 7, "- " + s)
        else:
            pdf.multi_cell(0, 7, suggestions)

        # Charts
        for title_k, buf in chart_buffers.items():
            try:
                pdf.add_page()
                pdf.cell(0, 8, title_k, ln=True)
                tmp_img = Path(tempfile.gettempdir()) / f"chart_{uuid.uuid4().hex}.png"
                with open(tmp_img, "wb") as f:
                    f.write(buf.getvalue())
                pdf.image(str(tmp_img), x=10, w=190)
                try:
                    os.remove(tmp_img)
                except Exception:
                    pass
            except Exception as e:
                # continue embedding other charts
                pdf.multi_cell(0, 6, f"[Could not embed chart {title_k}: {e}]")

        pdf.output(str(tmpname))
        with open(tmpname, "rb") as f:
            data = f.read()
        try:
            os.remove(tmpname)
        except Exception:
            pass
        return data
    except Exception as e:
        st.error(f"PDF generation failed: {e}")
        return None

# ---------------------------
# Load & clean input file
# ---------------------------
if uploaded is None:
    st.info("Please upload a CSV or Excel file containing the features (SYSTEM, EQUIPMENT, HEALTH INDICATOR (HI), SYNTHETIC VALUE, MIN–MAX THRESHOLDS, WORKING VALUE ONBOARD, REMARKS, Actual_Status). Use example dataset from the sidebar if you don't have one.")
    st.stop()

try:
    raw_bytes = uploaded.read()
    encoding = detect_encoding_bytes(raw_bytes)
    uploaded.seek(0)
    if uploaded.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded, encoding=encoding, on_bad_lines="skip")
    else:
        import openpyxl  # ensure available in environment
        df = pd.read_excel(uploaded, engine="openpyxl")
    df = sanitize_columns(df)
except Exception as e:
    st.error(f"Could not read uploaded file: {e}")
    st.stop()

st.subheader("Data preview (sanitized)")
st.dataframe(df.head(8))

# ---------------------------
# Identify expected columns (flexible)
# ---------------------------
minmax_col = find_column(df, ["MIN_MAX_THRESHOLDS", "MIN–MAX THRESHOLDS", "MIN MAX THRESHOLDS", "THRESHOLDS"])
working_col = find_column(df, ["WORKING_VALUE_ONBOARD", "WORKING VALUE ONBOARD", "WORKING_VALUE", "SYNTHETIC_VALUE", "SYNTHETIC VALUE"])
hi_indicator_col = find_column(df, ["HEALTH_INDICATOR_HI", "HEALTH INDICATOR", "HEALTH_INDICATOR", "HEALTH_INDICATOR_(HI)"])
actual_col = find_column(df, ["Actual_Status", "ACTUAL_STATUS", "STATUS", "CONDITION"])

if minmax_col is None:
    st.error("❌ Could not find a column like 'MIN–MAX THRESHOLDS' in the file. Please include it and re-upload.")
    st.stop()

# Extract thresholds & working values
df[["Min_Threshold", "Max_Threshold"]] = df[minmax_col].apply(lambda x: pd.Series(extract_min_max(x)))
if working_col and working_col in df.columns:
    df["Working_Value"] = df[working_col].apply(extract_numeric)
else:
    # fallback to SYNTHETIC_VALUE
    synth_col = find_column(df, ["SYNTHETIC_VALUE", "SYNTHETIC VALUE"])
    if synth_col and synth_col in df.columns:
        df["Working_Value"] = pd.to_numeric(df[synth_col], errors="coerce")
    else:
        df["Working_Value"] = np.nan

# normalize HI indicator column name if found
if hi_indicator_col and hi_indicator_col in df.columns:
    df = df.rename(columns={hi_indicator_col: "HEALTH_INDICATOR_HI"})
if actual_col and actual_col in df.columns:
    df = df.rename(columns={actual_col: "Actual_Status"})

# ---------------------------
# Rule-based evaluation
# ---------------------------
df["Health_Index"] = df.apply(lambda r: compute_hi_rule(r.get("Working_Value", np.nan), r.get("Min_Threshold", np.nan), r.get("Max_Threshold", np.nan)), axis=1)
df["Status_Rule"] = df["Health_Index"].apply(lambda v: ("Unknown" if pd.isna(v) else ("Healthy" if v==1.0 else ("Warning" if v==0.5 else "Critical"))))

# ---------------------------
# Visualizations (Rule-based)
# ---------------------------
sns.set_style("whitegrid")
chart_buffers = {}

# Status distribution pie
try:
    fig1, ax1 = plt.subplots(figsize=(6,6))
    df["Status_Rule"].value_counts().plot.pie(autopct="%1.1f%%", ax=ax1, startangle=90)
    ax1.set_ylabel("")
    ax1.set_title("Overall Condition Distribution (Rule-based)")
    chart_buffers["Status Distribution (Rule-based)"] = save_fig_to_buffer(fig1)
    st.pyplot(fig1)
except Exception as e:
    st.warning(f"Could not render status pie: {e}")

# HI distribution
try:
    fig2, ax2 = plt.subplots(figsize=(8,4))
    sns.histplot(df["Health_Index"].dropna(), kde=True, ax=ax2)
    ax2.set_title("Health Index Distribution (Rule-based)")
    chart_buffers["Health Index Distribution (Rule-based)"] = save_fig_to_buffer(fig2)
    st.pyplot(fig2)
except Exception as e:
    st.warning(f"Could not render HI distribution: {e}")

# HI by indicator (bar)
if "HEALTH_INDICATOR_HI" in df.columns:
    try:
        fig3, ax3 = plt.subplots(figsize=(12,6))
        sns.barplot(data=df, x="HEALTH_INDICATOR_HI", y="Health_Index", hue="Status_Rule", ax=ax3)
        ax3.set_xticklabels(ax3.get_xticklabels(), rotation=90)
        ax3.set_title("Health Index by Indicator (Rule-based)")
        chart_buffers["Health Index by Indicator (Rule-based)"] = save_fig_to_buffer(fig3)
        st.pyplot(fig3)
    except Exception as e:
        st.warning(f"Could not render HI by indicator: {e}")

# Average HI per system heatmap
if "SYSTEM" in df.columns:
    try:
        sys_hi = df.groupby("SYSTEM")["Health_Index"].mean().to_frame().fillna(0)
        fig4, ax4 = plt.subplots(figsize=(8, max(2, 0.5*len(sys_hi))))
        sns.heatmap(sys_hi, annot=True, cmap="YlGnBu", linewidths=0.5, ax=ax4)
        ax4.set_title("Average Health Index per System (Rule-based)")
        chart_buffers["Avg Health Index per System (Rule-based)"] = save_fig_to_buffer(fig4)
        st.pyplot(fig4)
    except Exception as e:
        st.warning(f"Could not render system heatmap: {e}")

# Working Value vs HI scatter
if "Working_Value" in df.columns:
    try:
        fig5, ax5 = plt.subplots(figsize=(8,5))
        sns.scatterplot(x="Working_Value", y="Health_Index", hue="Status_Rule", data=df, ax=ax5)
        ax5.set_title("Working Value vs Health Index (Rule-based)")
        chart_buffers["Working Value vs HI (Rule-based)"] = save_fig_to_buffer(fig5)
        st.pyplot(fig5)
    except Exception as e:
        st.warning(f"Could not render working vs HI scatter: {e}")

# HI by system boxplot
if "SYSTEM" in df.columns:
    try:
        fig6, ax6 = plt.subplots(figsize=(10,5))
        sns.boxplot(data=df, x="SYSTEM", y="Health_Index", ax=ax6)
        ax6.set_xticklabels(ax6.get_xticklabels(), rotation=45)
        ax6.set_title("Health Index by System (boxplot)")
        chart_buffers["HI by System (boxplot)"] = save_fig_to_buffer(fig6)
        st.pyplot(fig6)
    except Exception as e:
        st.warning(f"Could not render HI by system boxplot: {e}")

# ---------------------------
# Performance metrics summary for rule-based
# ---------------------------
total = len(df)
evaluated = int(df["Health_Index"].notna().sum())
coverage_pct = round(evaluated/total*100, 3) if total>0 else 0
hi_mean = round(df["Health_Index"].mean(skipna=True), 4) if evaluated>0 else np.nan
hi_std = round(df["Health_Index"].std(skipna=True), 4) if evaluated>0 else np.nan
hi_skew = round(df["Health_Index"].dropna().skew(), 4) if evaluated>0 else np.nan
hi_kurt = round(df["Health_Index"].dropna().kurtosis(), 4) if evaluated>0 else np.nan

metrics_df = pd.DataFrame({
    "Metric": ["Total Readings", "Evaluated Readings", "Coverage (%)", "Mean HI", "Std Dev HI", "Skewness", "Kurtosis"],
    "Value": [total, evaluated, coverage_pct, hi_mean if not pd.isna(hi_mean) else "N/A",
              hi_std if not pd.isna(hi_std) else "N/A",
              hi_skew if not pd.isna(hi_skew) else "N/A",
              hi_kurt if not pd.isna(hi_kurt) else "N/A"]
})
st.subheader("Performance Metrics (Rule-based)")
st.table(metrics_df)

# ---------------------------
# Machine Learning Prediction (if model exists and requested)
# ---------------------------
ml_df = None
ml_metrics = {}
ml_chart_buffers = {}
if model_bundle is not None and mode in ["Machine Learning (Random Forest)", "Compare both"]:
    st.subheader("Machine Learning (Random Forest) Predictions")
    try:
        feature_cols = model_bundle.get("feature_columns", ["Working_Value", "Min_Threshold", "Max_Threshold", "Health_Index"])
        # ensure Health_Index column exists
        if "Health_Index" not in df.columns:
            df["Health_Index"] = df.apply(lambda r: compute_hi_rule(r.get("Working_Value", np.nan), r.get("Min_Threshold", np.nan), r.get("Max_Threshold", np.nan)), axis=1)
        # Ensure DataFrame has all feature columns (fill missing with 0)
        X = df.reindex(columns=feature_cols).fillna(0)
        scaler = model_bundle.get("scaler", None)
        if scaler is not None:
            try:
                X_scaled = scaler.transform(X)
            except Exception as e:
                # try converting to numpy first
                X_scaled = scaler.transform(X.values)
        else:
            X_scaled = X.values

        model_obj = model_bundle["model"]
        preds = model_obj.predict(X_scaled)
        le = model_bundle.get("label_encoder", None)
        if le is not None:
            pred_labels = le.inverse_transform(preds)
        else:
            pred_labels = preds.astype(str)

        ml_df = df.copy()
        ml_df["Status_ML"] = pred_labels

        # prediction probabilities if available
        if hasattr(model_obj, "predict_proba"):
            try:
                probs = model_obj.predict_proba(X_scaled)
                top_idx = np.argmax(probs, axis=1)
                ml_df["Pred_Prob"] = probs[np.arange(len(probs)), top_idx]
            except Exception:
                pass

        st.dataframe(ml_df.head(10))

        # ML distribution chart
        figm, axm = plt.subplots(figsize=(6,4))
        ml_df["Status_ML"].value_counts().plot(kind="bar", ax=axm)
        axm.set_title("ML Predicted Status Distribution")
        ml_chart_buffers["ML Predicted Status"] = save_fig_to_buffer(figm)
        st.pyplot(figm)

        # Feature importance
        if hasattr(model_obj, "feature_importances_"):
            try:
                importances = model_obj.feature_importances_
                names = feature_cols
                idx = np.argsort(importances)[::-1]
                figfi, axfi = plt.subplots(figsize=(8,4))
                axfi.bar([names[i] for i in idx], importances[idx])
                axfi.set_title("Feature Importance (Random Forest)")
                axfi.set_xticklabels([names[i] for i in idx], rotation=45)
                ml_chart_buffers["Feature Importance (ML)"] = save_fig_to_buffer(figfi)
                st.pyplot(figfi)
            except Exception:
                pass

        # If Actual_Status present, compute classification metrics
        if "Actual_Status" in ml_df.columns and ml_df["Actual_Status"].notna().any():
            try:
                from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, precision_recall_fscore_support
                y_true = ml_df["Actual_Status"].astype(str).fillna("Unknown")
                y_pred = ml_df["Status_ML"].astype(str)
                acc = accuracy_score(y_true, y_pred)
                report = classification_report(y_true, y_pred, zero_division=0, output_dict=True)
                st.metric("ML Accuracy", f"{acc:.3f}")
                # show confusion matrix
                labels = sorted(list(set(y_true.unique()) | set(y_pred.unique())))
                cm = confusion_matrix(y_true, y_pred, labels=labels)
                figcm, axcm = plt.subplots(figsize=(5,4))
                sns.heatmap(cm, annot=True, fmt="d", ax=axcm, xticklabels=labels, yticklabels=labels, cmap="Blues")
                axcm.set_xlabel("Predicted"); axcm.set_ylabel("Actual"); axcm.set_title("Confusion Matrix (ML)")
                ml_chart_buffers["ML Confusion Matrix"] = save_fig_to_buffer(figcm)
                st.pyplot(figcm)

                # create a small metrics summary
                prec = np.mean([report[l]["precision"] for l in report if l not in ("accuracy","macro avg","weighted avg")])
                rec = np.mean([report[l]["recall"] for l in report if l not in ("accuracy","macro avg","weighted avg")])
                f1 = np.mean([report[l]["f1-score"] for l in report if l not in ("accuracy","macro avg","weighted avg")])
                ml_metrics.update({"accuracy": acc, "precision": prec, "recall": rec, "f1": f1})
            except Exception as e:
                st.warning(f"Could not compute ML metrics: {e}")

    except Exception as e:
        st.error(f"Error during ML prediction: {e}")

# ---------------------------
# Prepare unified chart buffers (rule + ml)
# ---------------------------
all_charts = {}
all_charts.update(chart_buffers)
all_charts.update(ml_chart_buffers)

# ---------------------------
# Interpretation & suggestions
# ---------------------------
if not pd.isna(hi_mean):
    if hi_mean >= 0.8:
        interp = "Overall fleet health is excellent."
        suggests = ["Continue routine inspections.", "Next inspection recommended in 60 days."]
    elif hi_mean >= 0.6:
        interp = "Fleet health is satisfactory — preventive maintenance advised."
        suggests = ["Schedule preventive maintenance in 30 days.", "Monitor Warning-tagged items closely."]
    else:
        interp = "Fleet health is poor — immediate maintenance required."
        suggests = ["Immediate inspection for Critical items.", "Prioritize critical systems for repair."]
else:
    interp = "Insufficient numeric data to compute Health Index."
    suggests = ["Provide numeric 'Working Value' and 'MIN–MAX THRESHOLDS' in the dataset."]

# ---------------------------
# Save results CSV / Download
# ---------------------------
st.subheader("Export / Reporting")

# Save combined results (with ML and rule columns)
out_df = df.copy()
if ml_df is not None:
    out_df["Status_ML"] = ml_df["Status_ML"].values

save_path = DATA_DIR / f"health_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
out_df.to_csv(save_path, index=False, encoding="utf-8")
with open(save_path, "rb") as f:
    st.download_button("📥 Download processed results (CSV)", data=f, file_name=save_path.name, mime="text/csv")

# Generate PDF button
if st.button("📘 Generate PDF Report"):
    title = "⚓ NAVAL MACHINERY HEALTH CONDITION REPORT"
    # Compose metrics for report: use metrics_df (rule-based) and ML metrics if present
    metrics_for_report = metrics_df.copy()
    if ml_metrics:
        # append ML metrics rows
        for k, v in ml_metrics.items():
            metrics_for_report = pd.concat([metrics_for_report, pd.DataFrame({"Metric":[f"ML_{k}"], "Value":[round(v,4)]})], ignore_index=True)
    pdf_bytes = generate_pdf_report(title, metrics_for_report, interp, suggests, all_charts)
    if pdf_bytes:
        st.download_button("📥 Download PDF Report", data=pdf_bytes, file_name="Naval_Machinery_Health_Report.pdf", mime="application/pdf")
    else:
        st.error("PDF generation failed. Check logs.")

# ---------------------------
# Maintenance log (append & download)
# ---------------------------
st.subheader("Maintenance Log")
log_entry = {
    "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "Mean_HI": float(hi_mean) if not pd.isna(hi_mean) else "",
    "Coverage_%": coverage_pct,
    "Mode": mode,
    "ML_Accuracy": round(ml_metrics.get("accuracy", np.nan), 4) if ml_metrics else ""
}
if LOG_PATH.exists():
    try:
        ldf = pd.read_csv(LOG_PATH)
        ldf = pd.concat([ldf, pd.DataFrame([log_entry])], ignore_index=True)
    except Exception:
        ldf = pd.DataFrame([log_entry])
else:
    ldf = pd.DataFrame([log_entry])
ldf.to_csv(LOG_PATH, index=False)
st.dataframe(ldf.tail(10))
with open(LOG_PATH, "rb") as f:
    st.download_button("📥 Download maintenance_log.csv", f, file_name="maintenance_log.csv", mime="text/csv")

# ---------------------------
# Final notes
# ---------------------------
st.markdown("---")
st.info("Done. Use the Download buttons to retrieve CSV/PDF reports. Add `src/rf_model.pkl` (bundle with model, scaler, label_encoder and feature_columns) to enable ML predictions. If you want me to tailor the PDF layout, export folder, or add a scheduler for periodic logging, say so.")
